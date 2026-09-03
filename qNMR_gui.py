import sys
import json
from pathlib import Path

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QDialog,
    QVBoxLayout, QHBoxLayout, QFormLayout,
    QLabel, QPushButton, QLineEdit, QDoubleSpinBox, QSpinBox,
    QGroupBox, QScrollArea, QSplitter, QTableWidget, QTableWidgetItem,
    QHeaderView, QDialogButtonBox, QMessageBox, QFileDialog, QFrame,
)
from PyQt6.QtCore import Qt, QUrl, QTimer, pyqtSignal, QLocale
from PyQt6.QtGui import QPixmap, QFont
from PyQt6.QtSvg import QSvgRenderer
from PyQt6.QtWebEngineWidgets import QWebEngineView
from PyQt6.QtWebEngineCore import QWebEngineSettings, QWebEngineScript

try:
    from rdkit import Chem
    from rdkit.Chem import Draw, Descriptors
    from PIL.ImageQt import ImageQt
    HAS_RDKIT = True
except ImportError:
    HAS_RDKIT = False

from qNMR import Compound, InternalStandard, StartingMaterial, QNMRCalculator, QNMRResult

def _base_dir() -> Path:
    # PyInstaller 6+ sets sys._MEIPASS for both onefile and onedir (_internal/)
    # Older PyInstaller onedir: use exe parent
    if getattr(sys, "frozen", False):
        return Path(getattr(sys, "_MEIPASS", str(Path(sys.executable).parent)))
    return Path(__file__).parent

RESOURCES = _base_dir() / "resources"
EDITOR_HTML = RESOURCES / "editor.html"
C_LOCALE = QLocale(QLocale.Language.C)


# ── Helpers ──────────────────────────────────────────────────────────────────

def _dspin(decimals=4, max_val=1e6, suffix="", step=None) -> QDoubleSpinBox:
    sb = QDoubleSpinBox()
    sb.setLocale(C_LOCALE)
    sb.setDecimals(decimals)
    sb.setRange(0.0, max_val)
    sb.setSingleStep(step if step else 10 ** (-decimals))
    if suffix:
        sb.setSuffix(suffix)
    return sb


def smiles_to_pixmap(smiles: str, w=190, h=145) -> QPixmap | None:
    if not HAS_RDKIT or not smiles:
        return None
    try:
        from rdkit.Chem.Draw import rdMolDraw2D
        from PyQt6.QtCore import QByteArray
        from PyQt6.QtGui import QPainter, QImage
        mol = Chem.MolFromSmiles(smiles)
        if mol is None:
            return None
        drawer = rdMolDraw2D.MolDraw2DSVG(w, h)
        drawer.DrawMolecule(mol)
        drawer.FinishDrawing()
        svg = drawer.GetDrawingText().encode("utf-8")
        renderer = QSvgRenderer(QByteArray(svg))
        img = QImage(w, h, QImage.Format.Format_ARGB32)
        img.fill(0xFFFFFFFF)
        painter = QPainter(img)
        renderer.render(painter)
        painter.end()
        return QPixmap.fromImage(img)
    except Exception:
        return None


def mw_from_smiles(smiles: str) -> float | None:
    if not HAS_RDKIT or not smiles:
        return None
    mol = Chem.MolFromSmiles(smiles)
    return Descriptors.MolWt(mol) if mol else None


# ── Molecule editor dialog ────────────────────────────────────────────────────

class MolEditorDialog(QDialog):
    def __init__(self, parent=None, initial_smiles: str = ""):
        super().__init__(parent)
        self.setWindowTitle("Draw / Edit Structure")
        self.resize(640, 580)
        self._smiles = initial_smiles
        self._initial = initial_smiles

        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(6)

        self.web = QWebEngineView()
        self.web.settings().setAttribute(
            QWebEngineSettings.WebAttribute.LocalContentCanAccessFileUrls, True
        )
        if EDITOR_HTML.exists():
            self.web.load(QUrl.fromLocalFile(str(EDITOR_HTML)))
        else:
            self.web.setHtml("<p style='padding:20px;font-family:sans-serif'>"
                             "Editor HTML not found at resources/editor.html</p>")
        self.web.loadFinished.connect(self._on_loaded)
        layout.addWidget(self.web)

        # Fallback SMILES text input
        row = QHBoxLayout()
        row.addWidget(QLabel("SMILES (fallback):"))
        self.smiles_edit = QLineEdit(initial_smiles)
        self.smiles_edit.setPlaceholderText("Paste SMILES here if the editor is unavailable")
        row.addWidget(self.smiles_edit)
        layout.addLayout(row)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.accepted.connect(self._on_ok)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _on_loaded(self, ok: bool):
        if ok and self._initial:
            self.web.page().runJavaScript(f"setSmiles({json.dumps(self._initial)})")

    def _on_ok(self):
        def got_smiles(js_smiles):
            smiles = (js_smiles or "").strip()
            if not smiles:
                smiles = self.smiles_edit.text().strip()
            if smiles and HAS_RDKIT:
                mol = Chem.MolFromSmiles(smiles)
                if mol:
                    smiles = Chem.MolToSmiles(mol)
            self._smiles = smiles
            # Defer accept() so the JS callback fully unwinds before the
            # QWebEngineView is destroyed alongside the dialog.
            QTimer.singleShot(0, self.accept)
        # PyQt6 requires the explicit worldId overload; passing callback as 2nd
        # arg without it causes an immediate type-conversion crash.
        self.web.page().runJavaScript(
            "getSmiles()",
            QWebEngineScript.ScriptWorldId.MainWorld,
            got_smiles,
        )

    @property
    def smiles(self) -> str:
        return self._smiles


# ── Compound input panel ──────────────────────────────────────────────────────

class CompoundPanel(QGroupBox):
    MODE_IS = "is"
    MODE_SM = "sm"
    MODE_COMPOUND = "compound"
    MODE_ANALYTE = "analyte"

    def __init__(self, title: str, mode: str = MODE_COMPOUND, parent=None):
        super().__init__(title, parent)
        self.mode = mode
        self._smiles = ""
        self._build()

    def _build(self):
        outer = QHBoxLayout(self)
        outer.setSpacing(12)

        # Left: molecule thumbnail + draw button
        left = QVBoxLayout()
        left.setSpacing(4)

        self.mol_label = QLabel("No structure")
        self.mol_label.setFixedSize(190, 145)
        self.mol_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.mol_label.setStyleSheet(
            "border: 1px solid #bbb; background: white; color: #999; font-size: 11px;"
        )
        left.addWidget(self.mol_label)

        draw_btn = QPushButton("Draw / Edit Structure")
        draw_btn.clicked.connect(self._on_draw)
        left.addWidget(draw_btn)
        left.addStretch()
        outer.addLayout(left)

        # Right: form
        form = QFormLayout()
        form.setSpacing(6)

        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText("compound name")
        form.addRow("Name:", self.name_edit)

        self.smiles_lbl = QLabel("<i>none</i>")
        self.smiles_lbl.setWordWrap(True)
        self.smiles_lbl.setMaximumWidth(240)
        form.addRow("SMILES:", self.smiles_lbl)

        self.mw_box = _dspin(3, 100_000, " g/mol", step=0.001)
        self.mw_box.valueChanged.connect(self._on_mw_changed)
        form.addRow("MW:", self.mw_box)

        if self.mode == self.MODE_IS:
            self.mass_box = _dspin(3, 1_000_000, " mg", step=0.1)
            self.mass_box.valueChanged.connect(self._sync_mass_to_mmol)
            form.addRow("Mass added:", self.mass_box)
            self.mmol_box = _dspin(5, 1_000_000, " mmol", step=0.00001)
            self.mmol_box.valueChanged.connect(self._sync_mmol_to_mass)
            form.addRow("Amount:", self.mmol_box)

        elif self.mode in (self.MODE_SM, self.MODE_ANALYTE):
            mass_lbl   = "Sample mass:"   if self.mode == self.MODE_ANALYTE else "Initial mass:"
            amount_lbl = "Sample amount:" if self.mode == self.MODE_ANALYTE else "Initial amount:"
            self.init_mass_box = _dspin(3, 1_000_000, " mg", step=0.1)
            self.init_mass_box.valueChanged.connect(self._sync_mass_to_mmol)
            form.addRow(mass_lbl, self.init_mass_box)
            self.init_mmol_box = _dspin(5, 1_000_000, " mmol", step=0.00001)
            self.init_mmol_box.valueChanged.connect(self._sync_mmol_to_mass)
            form.addRow(amount_lbl, self.init_mmol_box)

        self.nH_box = QSpinBox()
        self.nH_box.setLocale(C_LOCALE)
        self.nH_box.setRange(1, 999)
        self.nH_box.setValue(24 if self.mode == self.MODE_IS else 1)
        self.nH_box.setToolTip("Number of equivalent protons in the integrated NMR signal")
        form.addRow("Protons (N):", self.nH_box)

        self.area_box = _dspin(2, 1e10)
        if self.mode == self.MODE_IS:
            area_lbl = "Area (IS signal):"
        elif self.mode == self.MODE_SM:
            area_lbl = "Area (SM remaining):"
            self.area_box.setToolTip("Set to 0 if starting material is fully consumed")
        else:
            area_lbl = "NMR area:"
        form.addRow(area_lbl, self.area_box)

        outer.addLayout(form)

    # ── Mass ↔ mmol sync ──────────────────────────────────────────────────────

    def _on_mw_changed(self, _val=None):
        # When MW changes keep mass fixed, recompute mmol
        self._sync_mass_to_mmol()

    def _sync_mass_to_mmol(self, _val=None):
        mw = self.mw_box.value()
        if mw <= 0:
            return
        if self.mode == self.MODE_IS:
            mmol = self.mass_box.value() / mw   # mg / (g/mol) = mmol
            self.mmol_box.blockSignals(True)
            self.mmol_box.setValue(mmol)
            self.mmol_box.blockSignals(False)
        elif self.mode in (self.MODE_SM, self.MODE_ANALYTE):
            mmol = self.init_mass_box.value() / mw
            self.init_mmol_box.blockSignals(True)
            self.init_mmol_box.setValue(mmol)
            self.init_mmol_box.blockSignals(False)

    def _sync_mmol_to_mass(self, _val=None):
        mw = self.mw_box.value()
        if mw <= 0:
            return
        if self.mode == self.MODE_IS:
            mass_mg = self.mmol_box.value() * mw  # mmol * g/mol = mg
            self.mass_box.blockSignals(True)
            self.mass_box.setValue(mass_mg)
            self.mass_box.blockSignals(False)
        elif self.mode in (self.MODE_SM, self.MODE_ANALYTE):
            mass_mg = self.init_mmol_box.value() * mw
            self.init_mass_box.blockSignals(True)
            self.init_mass_box.setValue(mass_mg)
            self.init_mass_box.blockSignals(False)

    def _on_draw(self):
        dlg = MolEditorDialog(self, self._smiles)
        if dlg.exec() == QDialog.DialogCode.Accepted and dlg.smiles:
            self._apply_smiles(dlg.smiles)

    def _apply_smiles(self, smiles: str):
        try:
            self._smiles = smiles
            short = smiles[:38] + ("…" if len(smiles) > 38 else "")
            self.smiles_lbl.setText(short)
            self.smiles_lbl.setToolTip(smiles)

            px = smiles_to_pixmap(smiles, 190, 145)
            if px:
                self.mol_label.setPixmap(px)
                self.mol_label.setText("")

            mw = mw_from_smiles(smiles)
            if mw:
                self.mw_box.setValue(round(mw, 3))
                # valueChanged → _on_mw_changed → _sync_mass_to_mmol fires automatically

            if not self.name_edit.text() and len(smiles) <= 25:
                self.name_edit.setText(smiles)
        except Exception as e:
            print(f"[_apply_smiles] {e}")

    # ── serialization ─────────────────────────────────────────────────────────

    def get_data(self) -> dict:
        d = {
            "name": self.name_edit.text().strip(),
            "smiles": self._smiles,
            "mw": self.mw_box.value(),
            "n_protons": self.nH_box.value(),
            "area": self.area_box.value(),
        }
        if self.mode == self.MODE_IS:
            d["mass_mg"] = self.mass_box.value()
            d["mmol"] = self.mmol_box.value()
        if self.mode in (self.MODE_SM, self.MODE_ANALYTE):
            d["init_mass_mg"] = self.init_mass_box.value()
            d["init_mmol"] = self.init_mmol_box.value()
        return d

    def set_data(self, d: dict):
        if d.get("name"):
            self.name_edit.setText(d["name"])
        if d.get("smiles"):
            self._apply_smiles(d["smiles"])
        if d.get("mw"):
            self.mw_box.setValue(d["mw"])
        if d.get("n_protons"):
            self.nH_box.setValue(d["n_protons"])
        if "area" in d:
            self.area_box.setValue(d["area"])
        if self.mode == self.MODE_IS:
            if "mass_mg" in d:
                self.mass_box.setValue(d["mass_mg"])
            elif "mass" in d:           # old session saved grams
                self.mass_box.setValue(d["mass"] * 1000.0)
        if self.mode in (self.MODE_SM, self.MODE_ANALYTE):
            if "init_mass_mg" in d:
                self.init_mass_box.setValue(d["init_mass_mg"])
            elif "initial_mass" in d:   # old session saved grams
                self.init_mass_box.setValue(d["initial_mass"] * 1000.0)

    def get_compound(self):
        name = self.name_edit.text().strip() or "Unknown"
        mw = self.mw_box.value()
        n = self.nH_box.value()
        area = self.area_box.value()
        if self.mode == self.MODE_IS:
            return InternalStandard(
                name=name, smiles=self._smiles, mw=mw,
                n_protons=n, area=area,
                mass=self.mass_box.value() / 1000.0,   # mg → g
            )
        if self.mode in (self.MODE_SM, self.MODE_ANALYTE):
            return StartingMaterial(
                name=name, smiles=self._smiles, mw=mw,
                n_protons=n, area=area,
                initial_mass=self.init_mass_box.value() / 1000.0,  # mg → g
            )
        return Compound(name=name, smiles=self._smiles, mw=mw, n_protons=n, area=area)


# ── Removable compound panel (for products / byproducts) ─────────────────────

class RemovablePanel(QWidget):
    removed = pyqtSignal(object)

    def __init__(self, label: str, mode: str, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 4, 0, 0)
        layout.setSpacing(2)

        header = QHBoxLayout()
        header.addWidget(QLabel(f"<b>{label}</b>"))
        header.addStretch()
        rm = QPushButton("Remove")
        rm.setFixedWidth(70)
        rm.clicked.connect(lambda: self.removed.emit(self))
        header.addWidget(rm)
        layout.addLayout(header)

        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("color: #ccc;")
        layout.addWidget(sep)

        self.panel = CompoundPanel("", mode)
        layout.addWidget(self.panel)

    def get_compound(self):
        return self.panel.get_compound()

    def get_data(self) -> dict:
        return self.panel.get_data()

    def set_data(self, d: dict):
        self.panel.set_data(d)


# ── Purity check dialog ───────────────────────────────────────────────────────

class PurityDialog(QDialog):
    """
    qNMR purity determination.
    Formula: purity = (n_analyte × MW_analyte) / m_sample × 100 %
    where n_analyte is obtained from the standard qNMR equation.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Purity Check (qNMR)")
        self.setMinimumWidth(960)
        self.resize(980, 620)

        root = QVBoxLayout(self)
        root.setSpacing(10)

        # ── Two panels side by side ───────────────────────────────────────────
        panels = QHBoxLayout()
        panels.setSpacing(12)

        is_box = QGroupBox("Internal Standard")
        is_layout = QVBoxLayout(is_box)
        self.is_panel = CompoundPanel("", CompoundPanel.MODE_IS)
        is_layout.addWidget(self.is_panel)
        panels.addWidget(is_box)

        an_box = QGroupBox("Analyte (compound to check)")
        an_layout = QVBoxLayout(an_box)
        an_note = QLabel(
            "<i>Enter the total weighed mass of the sample "
            "(which may contain impurities) in 'Sample mass'.</i>"
        )
        an_note.setWordWrap(True)
        an_note.setStyleSheet("color:#555;font-size:11px;")
        an_layout.addWidget(an_note)
        self.an_panel = CompoundPanel("", CompoundPanel.MODE_ANALYTE)
        an_layout.addWidget(self.an_panel)
        panels.addWidget(an_box)

        root.addLayout(panels)

        # ── Bottom toolbar ────────────────────────────────────────────────────
        bottom = QHBoxLayout()

        calc_btn = QPushButton("Calculate Purity")
        calc_btn.setFixedHeight(32)
        calc_btn.setStyleSheet(
            "QPushButton{background:#7c3aed;color:white;border-radius:4px;}"
            "QPushButton:hover{background:#6d28d9;}"
        )
        calc_btn.clicked.connect(self._calculate)
        bottom.addWidget(calc_btn)

        bottom.addStretch()

        for text, slot in [("Save session", self._save), ("Load session", self._load)]:
            b = QPushButton(text)
            b.clicked.connect(slot)
            bottom.addWidget(b)

        root.addLayout(bottom)

        # ── Result label ──────────────────────────────────────────────────────
        self.result_lbl = QLabel("Enter values above and click Calculate.")
        self.result_lbl.setWordWrap(True)
        self.result_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.result_lbl.setStyleSheet(
            "background:#dcfce7;border:1px solid #4ade80;border-radius:6px;"
            "padding:12px;font-size:14px;color:#14532d;"
        )
        self.result_lbl.setMinimumHeight(52)
        root.addWidget(self.result_lbl)

        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        btns.rejected.connect(self.reject)
        root.addWidget(btns)

    def _calculate(self):
        try:
            std = self.is_panel.get_compound()

            if std.mw <= 0:
                raise ValueError("IS MW must be > 0")
            if std.mass <= 0:
                raise ValueError("IS mass must be > 0")
            if std.area <= 0:
                raise ValueError("IS NMR area must be > 0")

            an = self.an_panel.get_compound()
            sample_mass_g = an.initial_mass          # already in grams (mg/1000)
            if sample_mass_g <= 0:
                raise ValueError("Sample mass must be > 0")
            if an.mw <= 0:
                raise ValueError("Analyte MW must be > 0")
            if an.area <= 0:
                raise ValueError("Analyte NMR area must be > 0")

            n_is    = std.mass / std.mw              # mol
            n_an    = n_is * (an.area / std.area) * (std.n_protons / an.n_protons)
            m_pure  = n_an * an.mw                   # g  (mol × g/mol)
            purity  = m_pure / sample_mass_g * 100.0

            name = an.name if an.name != "Unknown" else "Analyte"
            self.result_lbl.setText(
                f"<b>n(IS)</b> = {n_is*1000:.4f} mmol"
                f"&nbsp;&nbsp;|&nbsp;&nbsp;"
                f"<b>n({name})</b> = {n_an*1000:.4f} mmol"
                f"&nbsp;&nbsp;|&nbsp;&nbsp;"
                f"<b>Pure mass</b> = {m_pure*1000:.2f} mg"
                f"&nbsp;&nbsp;|&nbsp;&nbsp;"
                f"<b style='font-size:15px;color:#166534;'>Purity = {purity:.1f} %</b>"
            )
        except Exception as exc:
            QMessageBox.warning(self, "Calculation error", str(exc))

    def _save(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Save purity session", str(Path.home()), "JSON files (*.json)"
        )
        if path:
            data = {"purity": True, "is": self.is_panel.get_data(), "analyte": self.an_panel.get_data()}
            Path(path).write_text(json.dumps(data, indent=2), encoding="utf-8")

    def _load(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Load purity session", str(Path.home()), "JSON files (*.json)"
        )
        if not path:
            return
        try:
            data = json.loads(Path(path).read_text(encoding="utf-8"))
            self.is_panel.set_data(data.get("is", {}))
            self.an_panel.set_data(data.get("analyte", {}))
        except Exception as exc:
            QMessageBox.warning(self, "Load error", str(exc))


# ── Main window ───────────────────────────────────────────────────────────────

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("qNMR Calculator")
        self.resize(1200, 780)
        self._product_panels: list[RemovablePanel] = []
        self._byproduct_panels: list[RemovablePanel] = []
        self._last_result: QNMRResult | None = None
        self._build()

    # ── UI construction ───────────────────────────────────────────────────────

    def _build(self):
        splitter = QSplitter(Qt.Orientation.Horizontal)
        self.setCentralWidget(splitter)

        # Left pane: scrollable input area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setMinimumWidth(540)
        scroll.setMaximumWidth(700)

        left_widget = QWidget()
        self.left_vbox = QVBoxLayout(left_widget)
        self.left_vbox.setContentsMargins(8, 8, 8, 8)
        self.left_vbox.setSpacing(10)

        self.is_panel = CompoundPanel("Internal Standard", CompoundPanel.MODE_IS)
        self.left_vbox.addWidget(self.is_panel)

        self.sm_panel = CompoundPanel("Starting Material", CompoundPanel.MODE_SM)
        self.left_vbox.addWidget(self.sm_panel)

        self.left_vbox.addLayout(
            self._section_header("Products", "Add Product", self._add_product)
        )
        self.products_vbox = QVBoxLayout()
        self.products_vbox.setSpacing(4)
        self.left_vbox.addLayout(self.products_vbox)

        self.left_vbox.addLayout(
            self._section_header("Byproducts", "Add Byproduct", self._add_byproduct)
        )
        self.byproducts_vbox = QVBoxLayout()
        self.byproducts_vbox.setSpacing(4)
        self.left_vbox.addLayout(self.byproducts_vbox)

        calc_btn = QPushButton("Calculate")
        calc_btn.setFixedHeight(44)
        calc_btn.setFont(QFont("", 13, QFont.Weight.Bold))
        calc_btn.setStyleSheet(
            "QPushButton{background:#2563eb;color:white;border-radius:6px;}"
            "QPushButton:hover{background:#1d4ed8;}"
        )
        calc_btn.clicked.connect(self._calculate)
        self.left_vbox.addWidget(calc_btn)
        self.left_vbox.addStretch()

        scroll.setWidget(left_widget)
        splitter.addWidget(scroll)

        # Right pane: results
        right_widget = QWidget()
        right_vbox = QVBoxLayout(right_widget)
        right_vbox.setContentsMargins(8, 8, 8, 8)
        right_vbox.setSpacing(8)

        toolbar = QHBoxLayout()
        toolbar.addWidget(QLabel("<b style='font-size:14px'>Results</b>"))
        toolbar.addStretch()
        purity_btn = QPushButton("Purity Check")
        purity_btn.setStyleSheet(
            "QPushButton{background:#7c3aed;color:white;border-radius:4px;padding:3px 8px;}"
            "QPushButton:hover{background:#6d28d9;}"
        )
        purity_btn.clicked.connect(self._open_purity)
        toolbar.addWidget(purity_btn)
        for text, slot in [
            ("Save session", self._save),
            ("Load session", self._load),
            ("Export Excel", self._export_excel),
        ]:
            btn = QPushButton(text)
            btn.clicked.connect(slot)
            toolbar.addWidget(btn)
        right_vbox.addLayout(toolbar)

        self.summary_lbl = QLabel("Run a calculation to see results.")
        self.summary_lbl.setWordWrap(True)
        self.summary_lbl.setStyleSheet(
            "background:#dbeafe;border:1px solid #3b82f6;border-radius:4px;"
            "padding:8px;font-size:13px;color:#1e3a5f;"
        )
        self.summary_lbl.setMinimumHeight(55)
        right_vbox.addWidget(self.summary_lbl)

        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(
            ["Compound", "Role", "n (mmol)", "Mass (mg)", "Yield (%)", "Selectivity (%)"]
        )
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        right_vbox.addWidget(self.table)

        splitter.addWidget(right_widget)
        splitter.setSizes([580, 620])

        file_menu = self.menuBar().addMenu("File")
        file_menu.addAction("Save session", self._save)
        file_menu.addAction("Load session", self._load)
        file_menu.addSeparator()
        file_menu.addAction("Export to Excel", self._export_excel)
        file_menu.addSeparator()
        file_menu.addAction("Quit", self.close)

        tools_menu = self.menuBar().addMenu("Tools")
        tools_menu.addAction("Purity Check…", self._open_purity)

    def _open_purity(self):
        dlg = PurityDialog(self)
        dlg.exec()

    def _section_header(self, title: str, btn_label: str, callback) -> QHBoxLayout:
        row = QHBoxLayout()
        lbl = QLabel(f"<b>{title}</b>")
        lbl.setStyleSheet("font-size:13px;")
        row.addWidget(lbl)
        row.addStretch()
        btn = QPushButton(f"+ {btn_label}")
        btn.setFixedWidth(130)
        btn.clicked.connect(callback)
        row.addWidget(btn)
        return row

    # ── Dynamic panels ────────────────────────────────────────────────────────

    def _add_product(self):
        idx = len(self._product_panels) + 1
        w = RemovablePanel(f"Product {idx}", CompoundPanel.MODE_COMPOUND)
        w.removed.connect(
            lambda x: self._remove_panel(x, self._product_panels, self.products_vbox)
        )
        self._product_panels.append(w)
        self.products_vbox.addWidget(w)

    def _add_byproduct(self):
        idx = len(self._byproduct_panels) + 1
        w = RemovablePanel(f"Byproduct {idx}", CompoundPanel.MODE_COMPOUND)
        w.removed.connect(
            lambda x: self._remove_panel(x, self._byproduct_panels, self.byproducts_vbox)
        )
        self._byproduct_panels.append(w)
        self.byproducts_vbox.addWidget(w)

    def _remove_panel(self, widget, panel_list: list, container: QVBoxLayout):
        if widget in panel_list:
            panel_list.remove(widget)
        container.removeWidget(widget)
        widget.deleteLater()

    # ── Calculation ───────────────────────────────────────────────────────────

    def _calculate(self):
        try:
            std = self.is_panel.get_compound()
            sm = self.sm_panel.get_compound()
            products = [w.get_compound() for w in self._product_panels]
            byproducts = [w.get_compound() for w in self._byproduct_panels]

            result = QNMRCalculator(std, sm, products, byproducts).calculate()
            self._last_result = result
            self._show_result(result)
        except Exception as exc:
            QMessageBox.warning(self, "Calculation error", str(exc))

    def _show_result(self, r: QNMRResult):
        parts = [f"<b>n(IS)</b> = {r.n_is * 1000:.4f} mmol"]
        if r.n_sm_initial > 0:
            parts.append(f"<b>n(SM initial)</b> = {r.n_sm_initial * 1000:.4f} mmol")
        if r.n_sm_remaining > 0:
            parts.append(f"<b>n(SM remaining)</b> = {r.n_sm_remaining * 1000:.4f} mmol")
        if r.conversion is not None:
            parts.append(f"<b>Conversion</b> = {r.conversion:.1f} %")
        self.summary_lbl.setText("  |  ".join(parts))

        all_entries = [(e, "Product") for e in r.products] + \
                      [(e, "Byproduct") for e in r.byproducts]
        self.table.setRowCount(len(all_entries))

        for row_i, (e, role) in enumerate(all_entries):
            self.table.setItem(row_i, 0, QTableWidgetItem(e["name"]))
            self.table.setItem(row_i, 1, QTableWidgetItem(role))
            self.table.setItem(row_i, 2, QTableWidgetItem(f"{e['moles'] * 1000:.4f}"))
            self.table.setItem(row_i, 3, QTableWidgetItem(f"{e['mass'] * 1000:.2f}"))
            yld = f"{e['yield']:.1f}" if e["yield"] is not None else "—"
            sel = f"{e['selectivity']:.1f}" if e["selectivity"] is not None else "—"
            self.table.setItem(row_i, 4, QTableWidgetItem(yld))
            self.table.setItem(row_i, 5, QTableWidgetItem(sel))

    # ── Session save/load ─────────────────────────────────────────────────────

    def _session_data(self) -> dict:
        return {
            "is": self.is_panel.get_data(),
            "sm": self.sm_panel.get_data(),
            "products": [w.get_data() for w in self._product_panels],
            "byproducts": [w.get_data() for w in self._byproduct_panels],
        }

    def _restore_session(self, data: dict):
        self.is_panel.set_data(data.get("is", {}))
        self.sm_panel.set_data(data.get("sm", {}))

        for w in list(self._product_panels):
            self._remove_panel(w, self._product_panels, self.products_vbox)
        for d in data.get("products", []):
            self._add_product()
            self._product_panels[-1].set_data(d)

        for w in list(self._byproduct_panels):
            self._remove_panel(w, self._byproduct_panels, self.byproducts_vbox)
        for d in data.get("byproducts", []):
            self._add_byproduct()
            self._byproduct_panels[-1].set_data(d)

    def _save(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Save session", str(Path.home()), "JSON files (*.json)"
        )
        if path:
            Path(path).write_text(json.dumps(self._session_data(), indent=2), encoding="utf-8")

    def _load(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Load session", str(Path.home()), "JSON files (*.json)"
        )
        if path:
            try:
                data = json.loads(Path(path).read_text(encoding="utf-8"))
                self._restore_session(data)
            except Exception as exc:
                QMessageBox.warning(self, "Load error", str(exc))

    # ── Excel export ──────────────────────────────────────────────────────────

    def _export_excel(self):
        if self._last_result is None:
            QMessageBox.information(self, "No results", "Run a calculation first.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Export to Excel", str(Path.home() / "qNMR_results.xlsx"),
            "Excel files (*.xlsx)"
        )
        if not path:
            return

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "qNMR Results"

        r = self._last_result
        blue_fill = PatternFill("solid", fgColor="2563EB")
        white_bold = Font(bold=True, color="FFFFFF")
        bold = Font(bold=True)

        ws["A1"] = "qNMR Calculator Results"
        ws["A1"].font = Font(bold=True, size=14)

        for row_i, (label, value) in enumerate([
            ("n(IS) [mmol]", round(r.n_is * 1000, 6)),
            ("n(SM initial) [mmol]", round(r.n_sm_initial * 1000, 6)),
            ("n(SM remaining) [mmol]", round(r.n_sm_remaining * 1000, 6)),
            ("Conversion [%]", round(r.conversion, 2) if r.conversion is not None else "N/A"),
        ], start=2):
            ws.cell(row=row_i, column=1, value=label).font = bold
            ws.cell(row=row_i, column=2, value=value)

        headers = ["Compound", "Role", "n [mmol]", "Mass [mg]", "Yield [%]", "Selectivity [%]"]
        for col_i, h in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_i, value=h)
            cell.font = white_bold
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal="center")

        row_i = 8
        for entry_list in (r.products, r.byproducts):
            role = "Product" if entry_list is r.products else "Byproduct"
            for e in entry_list:
                ws.cell(row=row_i, column=1, value=e["name"])
                ws.cell(row=row_i, column=2, value=role)
                ws.cell(row=row_i, column=3, value=round(e["moles"] * 1000, 6))
                ws.cell(row=row_i, column=4, value=round(e["mass"] * 1000, 4))
                ws.cell(row=row_i, column=5,
                        value=round(e["yield"], 2) if e["yield"] is not None else None)
                ws.cell(row=row_i, column=6,
                        value=round(e["selectivity"], 2) if e["selectivity"] is not None else None)
                row_i += 1

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

        wb.save(path)
        QMessageBox.information(self, "Exported", f"Saved to:\n{path}")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    QApplication.setAttribute(Qt.ApplicationAttribute.AA_ShareOpenGLContexts)
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    icon_path = _base_dir() / "qNMRicon.ico"
    if icon_path.exists():
        from PyQt6.QtGui import QIcon
        app.setWindowIcon(QIcon(str(icon_path)))
    win = MainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
